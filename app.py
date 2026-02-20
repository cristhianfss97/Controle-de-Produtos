from __future__ import annotations

import os
import io
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List

from flask import (
    Flask,
    flash,
    redirect,
    render_template_string,
    request,
    url_for,
    send_file,
    abort,
)
from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user,
    UserMixin,
)
from werkzeug.security import generate_password_hash, check_password_hash

from sqlalchemy import (
    Column,
    DateTime,
    ForeignKey,
    Integer,
    Numeric,
    String,
    Text,
    create_engine,
    select,
    func,
    text,
)
from sqlalchemy.orm import declarative_base, relationship, Session

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------------------------
# App
# ------------------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = os.environ.get("APP_SECRET_KEY", "dev-secret-change-me")

login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)


# ------------------------------------------------------------------------------
# Database
# ------------------------------------------------------------------------------

DATABASE_URL = (os.environ.get("DATABASE_URL") or "").strip()

# Render/Neon às vezes dão postgres:// (SQLAlchemy prefere postgresql://)
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

if DATABASE_URL:
    ENGINE = create_engine(DATABASE_URL, pool_pre_ping=True)
    USING_POSTGRES = True
else:
    ENGINE = create_engine("sqlite:///produtos.db", future=True)
    USING_POSTGRES = False

Base = declarative_base()


# ------------------------------------------------------------------------------
# Models
# ------------------------------------------------------------------------------

class Usuario(Base, UserMixin):
    __tablename__ = "usuarios"

    id = Column(Integer, primary_key=True, autoincrement=True)
    nome = Column(String(120), nullable=False)
    email = Column(String(180), nullable=False, unique=True, index=True)
    password_hash = Column(String(255), nullable=False)
    role = Column(String(20), nullable=False, default="operador")  # admin|operador
    ativo = Column(Integer, nullable=False, default=1)  # 1/0

    auditorias = relationship("Auditoria", back_populates="usuario")

    def get_id(self):
        return str(self.id)

    @property
    def is_active(self):
        return bool(self.ativo)


class Categoria(Base):
    __tablename__ = "categorias"

    id = Column(Integer, primary_key=True, autoincrement=True)
    nome = Column(String(120), nullable=False, unique=True, index=True)

    produtos = relationship("Produto", back_populates="categoria")


class Produto(Base):
    __tablename__ = "produtos"

    id = Column(Integer, primary_key=True, autoincrement=True)
    nome = Column(String(255), nullable=False, index=True)
    sku = Column(String(80), nullable=True, unique=True)
    preco = Column(Numeric(12, 2), nullable=False, default=0)
    quantidade = Column(Integer, nullable=False, default=0)
    estoque_minimo = Column(Integer, nullable=False, default=0)

    categoria_id = Column(Integer, ForeignKey("categorias.id", ondelete="SET NULL"), nullable=True)
    categoria = relationship("Categoria", back_populates="produtos")

    auditorias = relationship("Auditoria", back_populates="produto")


class Auditoria(Base):
    __tablename__ = "auditoria"

    id = Column(Integer, primary_key=True, autoincrement=True)
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)

    usuario_id = Column(Integer, ForeignKey("usuarios.id", ondelete="SET NULL"), nullable=True)
    usuario = relationship("Usuario", back_populates="auditorias")

    produto_id = Column(Integer, ForeignKey("produtos.id", ondelete="SET NULL"), nullable=True)
    produto = relationship("Produto", back_populates="auditorias")

    acao = Column(String(50), nullable=False)
    detalhe = Column(Text, nullable=True)
    ip = Column(String(64), nullable=True)


# ------------------------------------------------------------------------------
# DB init
# ------------------------------------------------------------------------------

def init_db():
    Base.metadata.create_all(ENGINE)
    with ENGINE.begin() as conn:
        # índice best-effort
        try:
            conn.execute(text("CREATE INDEX IF NOT EXISTS idx_audit_created ON auditoria (created_at)"))
        except Exception:
            pass

    # categoria default
    with Session(ENGINE) as s:
        if not s.execute(select(Categoria).limit(1)).scalar_one_or_none():
            s.add(Categoria(nome="Geral"))
            s.commit()


init_db()


# ------------------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------------------

def clean(s: str) -> str:
    return " ".join((s or "").strip().split())


def parse_int(s: str) -> Optional[int]:
    try:
        return int(str(s).strip())
    except Exception:
        return None


def parse_float(s: str) -> Optional[float]:
    try:
        return float(str(s).replace(",", ".").strip())
    except Exception:
        return None


def is_admin() -> bool:
    return bool(getattr(current_user, "role", "") == "admin")


def require_admin():
    if not is_admin():
        abort(403)


def get_ip() -> str:
    ip = request.headers.get("X-Forwarded-For", request.remote_addr) or ""
    return ip.split(",")[0].strip()[:64]


def log_event(acao: str, detalhe: str = "", produto_id: Optional[int] = None):
    uid = None
    if current_user and getattr(current_user, "is_authenticated", False):
        uid = int(current_user.id)
    with Session(ENGINE) as s:
        s.add(Auditoria(
            usuario_id=uid,
            produto_id=produto_id,
            acao=acao,
            detalhe=(detalhe or "")[:2000],
            ip=get_ip(),
        ))
        s.commit()


# ------------------------------------------------------------------------------
# Setup (primeiro admin) - só funciona se não existir usuário
# ------------------------------------------------------------------------------

def no_users_yet() -> bool:
    with Session(ENGINE) as s:
        total = s.execute(select(func.count(Usuario.id))).scalar_one()
        return total == 0


@app.before_request
def force_setup_if_empty():
    # impede 404 quebrar
    if request.endpoint is None:
        return
    if request.path.startswith("/static"):
        return
    if request.endpoint in {"setup", "setup_post"}:
        return

    if no_users_yet():
        return redirect(url_for("setup"))


SETUP_HTML = """
<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Primeiro acesso</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body { background: radial-gradient(1200px 800px at 20% -10%, #e9f2ff 0%, transparent 55%),
                   radial-gradient(1200px 800px at 95% 0%, #f4e9ff 0%, transparent 60%),
                   #f7f7fb; }
    .card { border:0; box-shadow:0 10px 30px rgba(0,0,0,.08); border-radius:18px; }
    .btn,.form-control{ border-radius:12px; }
  </style>
</head>
<body>
  <div class="container py-5" style="max-width:680px">
    <div class="card p-4">
      <h3 class="mb-1">Primeiro acesso</h3>
      <p class="text-muted mb-3">Crie o usuário <b>Administrador</b> da empresa. Depois disso, o cadastro público será desativado automaticamente.</p>

      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for category, message in messages %}
            <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
              {{ message }}
              <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
            </div>
          {% endfor %}
        {% endif %}
      {% endwith %}

      <form method="post" action="/setup">
        <div class="mb-2">
          <label class="form-label">Nome</label>
          <input class="form-control" name="nome" required>
        </div>
        <div class="mb-2">
          <label class="form-label">E-mail</label>
          <input class="form-control" type="email" name="email" required>
        </div>
        <div class="mb-3">
          <label class="form-label">Senha</label>
          <input class="form-control" type="password" name="senha" minlength="6" required>
          <div class="form-text">Mínimo 6 caracteres.</div>
        </div>
        <button class="btn btn-primary w-100" type="submit">Criar Admin</button>
      </form>
    </div>
  </div>
  <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""


@app.get("/setup")
def setup():
    if not no_users_yet():
        return redirect(url_for("login"))
    return render_template_string(SETUP_HTML)


@app.post("/setup")
def setup_post():
    if not no_users_yet():
        return redirect(url_for("login"))

    nome = clean(request.form.get("nome", ""))
    email = clean(request.form.get("email", "")).lower()
    senha = request.form.get("senha", "")

    if not nome or not email or len(senha) < 6:
        flash("Preencha corretamente (senha min. 6).", "danger")
        return redirect(url_for("setup"))

    with Session(ENGINE) as s:
        exists = s.execute(select(Usuario).where(func.lower(Usuario.email) == email).limit(1)).scalar_one_or_none()
        if exists:
            flash("Esse e-mail já está cadastrado.", "warning")
            return redirect(url_for("setup"))

        admin = Usuario(
            nome=nome,
            email=email,
            role="admin",
            password_hash=generate_password_hash(senha),
            ativo=1,
        )
        s.add(admin)
        s.commit()

    flash("Admin criado! Faça login.", "success")
    return redirect(url_for("login"))


# ------------------------------------------------------------------------------
# Login
# ------------------------------------------------------------------------------

@login_manager.user_loader
def load_user(user_id: str):
    with Session(ENGINE) as s:
        return s.get(Usuario, int(user_id))


LOGIN_HTML = """
<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Login</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body { background: radial-gradient(1200px 800px at 20% -10%, #e9f2ff 0%, transparent 55%),
                   radial-gradient(1200px 800px at 95% 0%, #f4e9ff 0%, transparent 60%),
                   #f7f7fb; }
    .card { border:0; box-shadow:0 10px 30px rgba(0,0,0,.08); border-radius:18px; }
    .btn,.form-control{ border-radius:12px; }
  </style>
</head>
<body>
  <div class="container py-5" style="max-width:520px">
    <div class="card p-4">
      <h3 class="mb-1">Entrar</h3>
      <p class="text-muted mb-3">Acesse o Controle de Produtos.</p>

      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for category, message in messages %}
            <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
              {{ message }}
              <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
            </div>
          {% endfor %}
        {% endif %}
      {% endwith %}

      <form method="post" action="/login">
        <div class="mb-2">
          <label class="form-label">E-mail</label>
          <input class="form-control" name="email" type="email" required>
        </div>
        <div class="mb-3">
          <label class="form-label">Senha</label>
          <input class="form-control" name="senha" type="password" required>
        </div>
        <button class="btn btn-primary w-100" type="submit">Entrar</button>
      </form>
    </div>
  </div>
  <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""


@app.get("/login")
def login():
    if no_users_yet():
        return redirect(url_for("setup"))
    if current_user.is_authenticated:
        return redirect(url_for("home"))
    return render_template_string(LOGIN_HTML)


@app.post("/login")
def login_post():
    if no_users_yet():
        return redirect(url_for("setup"))

    email = clean(request.form.get("email", "")).lower()
    senha = request.form.get("senha", "")

    with Session(ENGINE) as s:
        u = s.execute(select(Usuario).where(func.lower(Usuario.email) == email).limit(1)).scalar_one_or_none()

    if not u or not u.ativo or not check_password_hash(u.password_hash, senha):
        flash("Usuário/senha inválidos.", "danger")
        return redirect(url_for("login"))

    login_user(u)
    log_event("LOGIN", f"Login: {u.email}")
    return redirect(url_for("home"))


@app.get("/logout")
@login_required
def logout():
    log_event("LOGOUT", f"Logout: {current_user.email}")
    logout_user()
    return redirect(url_for("login"))


# ------------------------------------------------------------------------------
# UI Base
# ------------------------------------------------------------------------------

BASE_HTML = """
<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title or "Controle de Produtos" }}</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body { background: radial-gradient(1200px 800px at 20% -10%, #e9f2ff 0%, transparent 55%),
                   radial-gradient(1200px 800px at 95% 0%, #f4e9ff 0%, transparent 60%),
                   #f7f7fb; }
    .card { border:0; box-shadow:0 10px 30px rgba(0,0,0,.08); border-radius:18px; }
    .btn,.form-control,.form-select{ border-radius:12px; }
    .badge-soft{ background:rgba(13,110,253,.10); color:#0d6efd; border-radius:999px; padding:.25rem .55rem; }
    .pill{ background:rgba(13,110,253,.12); color:#0d6efd; border-radius:999px; padding:.28rem .6rem; font-weight:800; }
    .pill-warn{ background:rgba(255,193,7,.20); color:#8a6d00; }
    .mono{ font-variant-numeric: tabular-nums; }
    .muted{ color:rgba(0,0,0,.55); }
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg bg-white border-bottom sticky-top">
  <div class="container py-2">
    <a class="navbar-brand fw-semibold" href="{{ url_for('home') }}">📦 Controle de Produtos</a>
    <div class="d-flex gap-2 align-items-center flex-wrap justify-content-end">
      <span class="badge-soft small">{{ "Postgres" if using_postgres else "SQLite (local)" }}</span>
      <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('dashboard') }}">Dashboard</a>
      <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('auditoria') }}">Auditoria</a>
      {% if is_admin %}
        <a class="btn btn-sm btn-outline-primary" href="{{ url_for('usuarios') }}">Usuários</a>
      {% endif %}
      <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('export_produtos') }}">Exportar Produtos</a>
      <div class="ms-2 small text-secondary">
        {{ current_user.nome }} ({{ current_user.role }}) • <a href="{{ url_for('logout') }}">sair</a>
      </div>
    </div>
  </div>
</nav>

<main class="container py-4">
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for category, message in messages %}
        <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
          {{ message }}
          <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
        </div>
      {% endfor %}
    {% endif %}
  {% endwith %}
  {{ body|safe }}
</main>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

def render_page(body: str, **ctx):
    return render_template_string(
        BASE_HTML,
        body=render_template_string(body, **ctx),
        **ctx,
    )


# ------------------------------------------------------------------------------
# Home - Produtos
# ------------------------------------------------------------------------------

HOME_BODY = """
<div class="card">
  <div class="card-body">
    <div class="d-flex flex-wrap gap-3 align-items-center justify-content-between">
      <div>
        <h1 class="h4 mb-1">Produtos</h1>
        <div class="muted small">Pesquise antes de cadastrar (Nome ou SKU) para evitar duplicidade.</div>
      </div>

      <form class="d-flex gap-2 flex-wrap" method="get" action="{{ url_for('home') }}">
        <input class="form-control" style="min-width:220px" name="q" placeholder="Pesquisar por nome ou SKU..." value="{{ q }}">
        <select class="form-select" name="cat" style="min-width:200px">
          <option value="">Todas as categorias</option>
          {% for c in categorias %}
            <option value="{{ c.id }}" {% if cat_id and c.id==cat_id %}selected{% endif %}>{{ c.nome }}</option>
          {% endfor %}
        </select>
        <div class="form-check align-self-center">
          <input class="form-check-input" type="checkbox" name="low" value="1" id="low" {% if low_only %}checked{% endif %}>
          <label class="form-check-label" for="low">Somente estoque baixo</label>
        </div>
        <button class="btn btn-primary" type="submit">Pesquisar</button>
        <a class="btn btn-outline-secondary" href="{{ url_for('home') }}">Limpar</a>
      </form>
    </div>

    <hr>

    <div class="border rounded-4 p-3 bg-white" id="novo">
      <div class="d-flex justify-content-between align-items-center flex-wrap gap-2">
        <div class="fw-semibold">Adicionar produto</div>
        <div class="muted small">Preço aceita vírgula (10,50).</div>
      </div>
      <form method="post" action="{{ url_for('add_produto') }}" class="row g-2 mt-2">
        <div class="col-12 col-lg-4">
          <input class="form-control" name="nome" placeholder="Nome do produto" required>
        </div>
        <div class="col-12 col-lg-2">
          <input class="form-control" name="sku" placeholder="SKU (opcional)">
        </div>
        <div class="col-6 col-lg-2">
          <div class="input-group">
            <span class="input-group-text">R$</span>
            <input class="form-control" name="preco" placeholder="199.90" required>
          </div>
        </div>
        <div class="col-6 col-lg-1">
          <input class="form-control" name="quantidade" placeholder="Qtd" required>
        </div>
        <div class="col-6 col-lg-1">
          <input class="form-control" name="minimo" placeholder="Mín" value="0" required>
        </div>
        <div class="col-6 col-lg-2">
          <select class="form-select" name="categoria_id">
            <option value="">Sem categoria</option>
            {% for c in categorias %}
              <option value="{{ c.id }}">{{ c.nome }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-12 d-grid">
          <button class="btn btn-success" type="submit">Salvar</button>
        </div>
      </form>

      <div class="d-flex justify-content-between align-items-center mt-3">
        <div class="muted small"><b>Estoque baixo</b> quando <span class="mono">Qtd ≤ Mín</span>.</div>
        <button class="btn btn-sm btn-outline-primary" data-bs-toggle="modal" data-bs-target="#catsModal">Categorias</button>
      </div>
    </div>

    <hr>

    <div class="table-responsive">
      <table class="table table-hover bg-white rounded-4 overflow-hidden">
        <thead class="table-light">
          <tr>
            <th>ID</th>
            <th>Produto</th>
            <th>Categoria</th>
            <th>SKU</th>
            <th>Preço</th>
            <th>Estoque</th>
            <th class="text-end">Ações</th>
          </tr>
        </thead>
        <tbody>
          {% if not produtos %}
            <tr><td colspan="7" class="text-center text-secondary py-5">Nenhum produto encontrado.</td></tr>
          {% endif %}

          {% for p in produtos %}
            {% set low = (p.quantidade <= p.estoque_minimo) %}
            <tr>
              <td class="text-secondary">#{{ p.id }}</td>
              <td>
                <div class="fw-semibold">{{ p.nome }}</div>
                <div class="muted small">Mín: <span class="mono">{{ p.estoque_minimo }}</span></div>
              </td>
              <td class="muted">{{ p.categoria.nome if p.categoria else "—" }}</td>
              <td class="mono">{{ p.sku if p.sku else "—" }}</td>
              <td class="mono">R$ {{ '%.2f'|format(p.preco) }}</td>
              <td>
                <span class="pill mono {% if low %}pill-warn{% endif %}">
                  {{ p.quantidade }}{% if low %} • baixo{% endif %}
                </span>
              </td>
              <td class="text-end">
                <div class="d-inline-flex flex-wrap gap-2 justify-content-end">

                  <form method="post" action="{{ url_for('stock_delta', pid=p.id) }}">
                    <input type="hidden" name="delta" value="1">
                    <input type="hidden" name="motivo" value="Ajuste rápido +1">
                    <button class="btn btn-sm btn-outline-success" type="submit">+1</button>
                  </form>

                  <form method="post" action="{{ url_for('stock_delta', pid=p.id) }}">
                    <input type="hidden" name="delta" value="-1">
                    <input type="hidden" name="motivo" value="Ajuste rápido -1">
                    <button class="btn btn-sm btn-outline-warning" type="submit">-1</button>
                  </form>

                  <button class="btn btn-sm btn-outline-secondary"
                          data-bs-toggle="modal"
                          data-bs-target="#stockModal"
                          data-id="{{ p.id }}"
                          data-nome="{{ p.nome|e }}"
                          data-qtd="{{ p.quantidade }}">
                    Ajustar
                  </button>

                  <form method="post" action="{{ url_for('delete_produto', pid=p.id) }}"
                        onsubmit="return confirm('Excluir o produto: {{ p.nome }} ?');">
                    <button class="btn btn-sm btn-outline-danger" type="submit">Excluir</button>
                  </form>
                </div>
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>

      <div class="muted small">Mostrando: <b>{{ produtos|length }}</b> • Total: <b>{{ total }}</b></div>
    </div>
  </div>
</div>

<!-- Modal Ajustar Estoque -->
<div class="modal fade" id="stockModal" tabindex="-1" aria-hidden="true">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content" style="border-radius: 18px;">
      <div class="modal-header">
        <h5 class="modal-title">Ajustar estoque</h5>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>

      <form method="post" id="stock-form">
        <div class="modal-body">
          <div class="mb-2">
            <div class="fw-semibold" id="stock-nome"></div>
            <div class="muted small">Quantidade atual: <span class="mono" id="stock-atual"></span></div>
          </div>
          <div class="mb-2">
            <label class="form-label">Nova quantidade</label>
            <input class="form-control" name="quantidade_nova" id="stock-nova" required>
          </div>
          <div class="mb-2">
            <label class="form-label">Motivo</label>
            <input class="form-control" name="motivo" placeholder="Ex.: Contagem / Entrada / Saída" required>
          </div>
          <div class="form-text">Esse ajuste gera registro na Auditoria.</div>
        </div>
        <div class="modal-footer">
          <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Cancelar</button>
          <button type="submit" class="btn btn-primary">Salvar</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Modal Categorias -->
<div class="modal fade" id="catsModal" tabindex="-1" aria-hidden="true">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content" style="border-radius: 18px;">
      <div class="modal-header">
        <h5 class="modal-title">Categorias</h5>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <form class="d-flex gap-2" method="post" action="{{ url_for('add_categoria') }}">
          <input class="form-control" name="nome" placeholder="Nova categoria" required>
          <button class="btn btn-primary" type="submit">Adicionar</button>
        </form>
        <hr>
        <ul class="list-group">
          {% for c in categorias %}
            <li class="list-group-item d-flex justify-content-between align-items-center">
              <span>{{ c.nome }}</span>
              <form method="post" action="{{ url_for('delete_categoria', cid=c.id) }}"
                    onsubmit="return confirm('Excluir a categoria {{ c.nome }}? Produtos ficarão sem categoria.');">
                <button class="btn btn-sm btn-outline-danger" type="submit">Excluir</button>
              </form>
            </li>
          {% endfor %}
        </ul>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Fechar</button>
      </div>
    </div>
  </div>
</div>

<script>
  const stockModal = document.getElementById('stockModal');
  stockModal?.addEventListener('show.bs.modal', event => {
    const btn = event.relatedTarget;
    const pid = btn.getAttribute('data-id');
    const nome = btn.getAttribute('data-nome');
    const qtd = btn.getAttribute('data-qtd');
    document.getElementById('stock-nome').innerText = nome;
    document.getElementById('stock-atual').innerText = qtd;
    document.getElementById('stock-nova').value = qtd;
    document.getElementById('stock-form').action = "/stock/set/" + pid;
  });
</script>
"""

@app.get("/")
@login_required
def home():
    q = (request.args.get("q") or "").strip()
    cat = (request.args.get("cat") or "").strip()
    low_only = (request.args.get("low") or "").strip() == "1"

    cat_id = None
    if cat:
        try:
            cat_id = int(cat)
        except Exception:
            cat_id = None

    with Session(ENGINE) as s:
        categorias = s.execute(select(Categoria).order_by(func.lower(Categoria.nome))).scalars().all()
        total = s.execute(select(func.count(Produto.id))).scalar_one()

        stmt = select(Produto).order_by(Produto.id.desc())

        if q:
            ql = q.lower()
            stmt = stmt.where(
                func.lower(Produto.nome).like(f"%{ql}%")
                | func.lower(func.coalesce(Produto.sku, "")).like(f"%{ql}%")
            )
        if cat_id:
            stmt = stmt.where(Produto.categoria_id == cat_id)
        if low_only:
            stmt = stmt.where(Produto.quantidade <= Produto.estoque_minimo)

        produtos = s.execute(stmt).scalars().all()
        for p in produtos:
            _ = p.categoria

    return render_page(
        HOME_BODY,
        title="Produtos",
        produtos=produtos,
        categorias=categorias,
        total=total,
        q=q,
        cat_id=cat_id,
        low_only=low_only,
        using_postgres=USING_POSTGRES,
        is_admin=is_admin(),
        current_user=current_user,
    )


@app.post("/add")
@login_required
def add_produto():
    nome = clean(request.form.get("nome", ""))
    sku = (request.form.get("sku", "") or "").strip()
    preco = parse_float(request.form.get("preco", ""))
    qtd = parse_int(request.form.get("quantidade", ""))
    minimo = parse_int(request.form.get("minimo", "0"))
    cat_raw = (request.form.get("categoria_id") or "").strip()
    cat_id = int(cat_raw) if cat_raw.isdigit() else None

    if not nome:
        flash("Nome obrigatório.", "danger")
        return redirect(url_for("home") + "#novo")
    if preco is None:
        flash("Preço inválido.", "danger")
        return redirect(url_for("home") + "#novo")
    if qtd is None or qtd < 0:
        flash("Quantidade inválida.", "danger")
        return redirect(url_for("home") + "#novo")
    if minimo is None or minimo < 0:
        flash("Estoque mínimo inválido.", "danger")
        return redirect(url_for("home") + "#novo")

    with Session(ENGINE) as s:
        if sku:
            dup = s.execute(select(Produto).where(func.lower(Produto.sku) == sku.lower()).limit(1)).scalar_one_or_none()
            if dup:
                flash(f"SKU já existe (ID #{dup.id}).", "warning")
                return redirect(url_for("home", q=sku))

        dup_nome = s.execute(select(Produto).where(func.lower(Produto.nome) == nome.lower()).limit(1)).scalar_one_or_none()
        if dup_nome:
            flash(f"Já existe produto com esse nome (ID #{dup_nome.id}).", "warning")
            return redirect(url_for("home", q=nome))

        p = Produto(
            nome=nome,
            sku=sku or None,
            preco=preco,
            quantidade=qtd,
            estoque_minimo=minimo,
            categoria_id=cat_id,
        )
        s.add(p)
        s.commit()

        log_event("CREATE_PRODUCT", f"Criou produto #{p.id}: {p.nome} (SKU={p.sku})", produto_id=p.id)

    flash("Produto cadastrado!", "success")
    return redirect(url_for("home"))


@app.post("/delete/<int:pid>")
@login_required
def delete_produto(pid: int):
    with Session(ENGINE) as s:
        p = s.get(Produto, pid)
        if not p:
            flash("Produto não encontrado.", "warning")
            return redirect(url_for("home"))
        nome = p.nome
        s.delete(p)
        s.commit()
        log_event("DELETE_PRODUCT", f"Excluiu produto #{pid}: {nome}", produto_id=pid)

    flash("Produto excluído.", "success")
    return redirect(url_for("home"))


@app.post("/stock/delta/<int:pid>")
@login_required
def stock_delta(pid: int):
    delta = parse_int(request.form.get("delta", "0"))
    motivo = clean(request.form.get("motivo", ""))[:250]

    if delta is None:
        flash("Ação inválida.", "danger")
        return redirect(url_for("home"))

    with Session(ENGINE) as s:
        p = s.get(Produto, pid)
        if not p:
            flash("Produto não encontrado.", "warning")
            return redirect(url_for("home"))

        antes = int(p.quantidade)
        novo = antes + int(delta)
        if novo < 0:
            flash("Estoque não pode ficar negativo.", "warning")
            return redirect(url_for("home"))

        p.quantidade = novo
        s.commit()

        log_event("UPDATE_STOCK", f"Produto #{p.id} {p.nome}: {antes}->{novo}. Motivo: {motivo}", produto_id=p.id)

    flash("Estoque atualizado!", "success")
    return redirect(url_for("home"))


@app.post("/stock/set/<int:pid>")
@login_required
def stock_set(pid: int):
    novo = parse_int(request.form.get("quantidade_nova", ""))
    motivo = clean(request.form.get("motivo", ""))[:250]

    if novo is None or novo < 0 or not motivo:
        flash("Dados inválidos (quantidade >= 0 e motivo obrigatório).", "danger")
        return redirect(url_for("home"))

    with Session(ENGINE) as s:
        p = s.get(Produto, pid)
        if not p:
            flash("Produto não encontrado.", "warning")
            return redirect(url_for("home"))

        antes = int(p.quantidade)
        p.quantidade = novo
        s.commit()

        log_event("UPDATE_STOCK", f"Produto #{p.id} {p.nome}: {antes}->{novo}. Motivo: {motivo}", produto_id=p.id)

    flash("Ajuste registrado!", "success")
    return redirect(url_for("home"))


@app.post("/categorias/add")
@login_required
def add_categoria():
    nome = clean(request.form.get("nome", ""))
    if not nome:
        flash("Nome da categoria obrigatório.", "danger")
        return redirect(url_for("home"))

    with Session(ENGINE) as s:
        dup = s.execute(select(Categoria).where(func.lower(Categoria.nome) == nome.lower()).limit(1)).scalar_one_or_none()
        if dup:
            flash("Categoria já existe.", "warning")
            return redirect(url_for("home"))
        s.add(Categoria(nome=nome))
        s.commit()

    log_event("CREATE_CATEGORY", f"Criou categoria: {nome}")
    flash("Categoria criada!", "success")
    return redirect(url_for("home"))


@app.post("/categorias/delete/<int:cid>")
@login_required
def delete_categoria(cid: int):
    with Session(ENGINE) as s:
        c = s.get(Categoria, cid)
        if not c:
            flash("Categoria não encontrada.", "warning")
            return redirect(url_for("home"))

        prods = s.execute(select(Produto).where(Produto.categoria_id == cid)).scalars().all()
        for p in prods:
            p.categoria_id = None
        nome = c.nome
        s.delete(c)
        s.commit()

    log_event("DELETE_CATEGORY", f"Excluiu categoria: {nome} (afetou {len(prods)} produtos)")
    flash("Categoria excluída.", "success")
    return redirect(url_for("home"))


# ------------------------------------------------------------------------------
# Admin - Usuários
# ------------------------------------------------------------------------------

USERS_BODY = """
<div class="card">
  <div class="card-body">
    <div class="d-flex justify-content-between align-items-center flex-wrap gap-2">
      <div>
        <h1 class="h4 mb-1">Usuários</h1>
        <div class="muted small">Somente Admin pode criar/ativar/desativar usuários.</div>
      </div>
      <a class="btn btn-outline-secondary" href="{{ url_for('home') }}">Voltar</a>
    </div>

    <hr>

    <form class="row g-2" method="post" action="{{ url_for('create_user') }}">
      <div class="col-12 col-lg-3"><input class="form-control" name="nome" placeholder="Nome" required></div>
      <div class="col-12 col-lg-3"><input class="form-control" type="email" name="email" placeholder="E-mail" required></div>
      <div class="col-12 col-lg-2"><input class="form-control" name="senha" placeholder="Senha (min 6)" required></div>
      <div class="col-12 col-lg-2">
        <select class="form-select" name="role">
          <option value="operador">Operador</option>
          <option value="admin">Admin</option>
        </select>
      </div>
      <div class="col-12 col-lg-2 d-grid"><button class="btn btn-primary" type="submit">Criar</button></div>
    </form>

    <hr>

    <div class="table-responsive">
      <table class="table table-hover bg-white rounded-4 overflow-hidden">
        <thead class="table-light">
          <tr>
            <th>ID</th><th>Nome</th><th>E-mail</th><th>Role</th><th>Status</th><th class="text-end">Ações</th>
          </tr>
        </thead>
        <tbody>
          {% for u in usuarios %}
            <tr>
              <td class="text-secondary">#{{ u.id }}</td>
              <td>{{ u.nome }}</td>
              <td class="mono">{{ u.email }}</td>
              <td><span class="badge-soft">{{ u.role }}</span></td>
              <td>{{ "Ativo" if u.ativo else "Inativo" }}</td>
              <td class="text-end">
                {% if u.id != current_user.id %}
                  <form method="post" class="d-inline" action="{{ url_for('toggle_user', uid=u.id) }}">
                    <button class="btn btn-sm btn-outline-warning" type="submit">{{ "Desativar" if u.ativo else "Ativar" }}</button>
                  </form>
                {% else %}
                  <span class="text-muted small">você</span>
                {% endif %}
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
</div>
"""

@app.get("/usuarios")
@login_required
def usuarios():
    require_admin()
    with Session(ENGINE) as s:
        usuarios = s.execute(select(Usuario).order_by(Usuario.id.asc())).scalars().all()
    return render_page(
        USERS_BODY,
        title="Usuários",
        usuarios=usuarios,
        using_postgres=USING_POSTGRES,
        is_admin=is_admin(),
        current_user=current_user,
    )


@app.post("/usuarios/create")
@login_required
def create_user():
    require_admin()

    nome = clean(request.form.get("nome", ""))
    email = clean(request.form.get("email", "")).lower()
    senha = request.form.get("senha", "")
    role = (request.form.get("role") or "operador").strip()
    if role not in {"admin", "operador"}:
        role = "operador"

    if not nome or not email or len(senha) < 6:
        flash("Preencha nome, e-mail e senha (min 6).", "danger")
        return redirect(url_for("usuarios"))

    with Session(ENGINE) as s:
        exists = s.execute(select(Usuario).where(func.lower(Usuario.email) == email).limit(1)).scalar_one_or_none()
        if exists:
            flash("E-mail já cadastrado.", "warning")
            return redirect(url_for("usuarios"))

        u = Usuario(
            nome=nome,
            email=email,
            role=role,
            password_hash=generate_password_hash(senha),
            ativo=1,
        )
        s.add(u)
        s.commit()

    log_event("CREATE_USER", f"Criou usuário: {email} role={role}")
    flash("Usuário criado!", "success")
    return redirect(url_for("usuarios"))


@app.post("/usuarios/toggle/<int:uid>")
@login_required
def toggle_user(uid: int):
    require_admin()
    if int(uid) == int(current_user.id):
        flash("Você não pode desativar a si mesmo.", "warning")
        return redirect(url_for("usuarios"))

    with Session(ENGINE) as s:
        u = s.get(Usuario, uid)
        if not u:
            flash("Usuário não encontrado.", "warning")
            return redirect(url_for("usuarios"))
        u.ativo = 0 if u.ativo else 1
        s.commit()

    log_event("TOGGLE_USER", f"Alterou status: {u.email} ativo={u.ativo}")
    flash("Status atualizado.", "success")
    return redirect(url_for("usuarios"))


# ------------------------------------------------------------------------------
# Auditoria
# ------------------------------------------------------------------------------

AUDIT_BODY = """
<div class="card">
  <div class="card-body">
    <div class="d-flex justify-content-between align-items-center flex-wrap gap-2">
      <div>
        <h1 class="h4 mb-1">Auditoria</h1>
        <div class="muted small">Histórico de ações e movimentações (últimos 200).</div>
      </div>
      <div class="d-flex gap-2">
        <a class="btn btn-outline-secondary" href="{{ url_for('export_auditoria') }}">Exportar Excel</a>
        <a class="btn btn-outline-secondary" href="{{ url_for('home') }}">Voltar</a>
      </div>
    </div>

    <hr>

    <form class="row g-2" method="get" action="{{ url_for('auditoria') }}">
      <div class="col-12 col-lg-4">
        <input class="form-control" name="q" placeholder="Buscar (ação, detalhe, email...)" value="{{ q }}">
      </div>
      <div class="col-6 col-lg-3">
        <input class="form-control" name="acao" placeholder="Ação (ex: UPDATE_STOCK)" value="{{ acao }}">
      </div>
      <div class="col-6 col-lg-2 d-grid">
        <button class="btn btn-primary" type="submit">Filtrar</button>
      </div>
      <div class="col-12 col-lg-3 d-grid">
        <a class="btn btn-outline-secondary" href="{{ url_for('auditoria') }}">Limpar</a>
      </div>
    </form>

    <hr>

    <div class="table-responsive">
      <table class="table table-hover bg-white rounded-4 overflow-hidden">
        <thead class="table-light">
          <tr>
            <th>Data</th><th>Ação</th><th>Usuário</th><th>Produto</th><th>Detalhe</th><th>IP</th>
          </tr>
        </thead>
        <tbody>
          {% for a in itens %}
            <tr>
              <td class="mono text-secondary">{{ a.created_at.strftime("%Y-%m-%d %H:%M:%S") }}</td>
              <td><span class="badge-soft">{{ a.acao }}</span></td>
              <td class="mono">{{ a.usuario.email if a.usuario else "—" }}</td>
              <td class="mono">{{ a.produto_id if a.produto_id else "—" }}</td>
              <td class="small">{{ a.detalhe or "" }}</td>
              <td class="mono">{{ a.ip or "—" }}</td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

  </div>
</div>
"""

@app.get("/auditoria")
@login_required
def auditoria():
    q = clean(request.args.get("q", ""))
    acao = clean(request.args.get("acao", ""))

    with Session(ENGINE) as s:
        stmt = select(Auditoria).order_by(Auditoria.id.desc()).limit(200)

        if acao:
            stmt = stmt.where(Auditoria.acao == acao)

        if q:
            ql = f"%{q.lower()}%"
            stmt = stmt.where(
                func.lower(func.coalesce(Auditoria.acao, "")).like(ql)
                | func.lower(func.coalesce(Auditoria.detalhe, "")).like(ql)
                | func.lower(func.coalesce(Auditoria.ip, "")).like(ql)
            )

        itens = s.execute(stmt).scalars().all()
        for a in itens:
            _ = a.usuario
            _ = a.produto

    return render_page(
        AUDIT_BODY,
        title="Auditoria",
        itens=itens,
        q=q,
        acao=acao,
        using_postgres=USING_POSTGRES,
        is_admin=is_admin(),
        current_user=current_user,
    )


# ------------------------------------------------------------------------------
# Dashboard
# ------------------------------------------------------------------------------

DASH_BODY = """
<div class="row g-3">
  <div class="col-12 col-lg-4">
    <div class="card"><div class="card-body">
      <div class="muted small">Total de produtos</div>
      <div class="display-6 fw-bold mono">{{ kpi_total_produtos }}</div>
    </div></div>
  </div>
  <div class="col-12 col-lg-4">
    <div class="card"><div class="card-body">
      <div class="muted small">Estoque baixo</div>
      <div class="display-6 fw-bold mono">{{ kpi_low }}</div>
    </div></div>
  </div>
  <div class="col-12 col-lg-4">
    <div class="card"><div class="card-body">
      <div class="muted small">Movimentações (7 dias)</div>
      <div class="display-6 fw-bold mono">{{ kpi_moves7 }}</div>
    </div></div>
  </div>

  <div class="col-12">
    <div class="card"><div class="card-body">
      <div class="d-flex justify-content-between align-items-center flex-wrap gap-2">
        <div>
          <h1 class="h5 mb-1">Movimentações por dia</h1>
          <div class="muted small">Últimos 7 dias (ações UPDATE_STOCK).</div>
        </div>
        <a class="btn btn-outline-secondary" href="{{ url_for('home') }}">Voltar</a>
      </div>

      <hr>
      <canvas id="chart"></canvas>
    </div></div>
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
  const labels = {{ labels|tojson }};
  const values = {{ values|tojson }};
  const ctx = document.getElementById('chart');

  new Chart(ctx, {
    type: 'line',
    data: {
      labels: labels,
      datasets: [{ label: 'Movimentações', data: values, tension: 0.25 }]
    },
    options: {
      responsive: true,
      plugins: { legend: { display: true } },
      scales: { y: { beginAtZero: true } }
    }
  });
</script>
"""

@app.get("/dashboard")
@login_required
def dashboard():
    today = datetime.utcnow().date()
    days = [(today - timedelta(days=i)) for i in range(6, -1, -1)]
    labels = [d.strftime("%d/%m") for d in days]
    values = [0] * len(days)

    with Session(ENGINE) as s:
        kpi_total = s.execute(select(func.count(Produto.id))).scalar_one()
        kpi_low = s.execute(select(func.count(Produto.id)).where(Produto.quantidade <= Produto.estoque_minimo)).scalar_one()

        # Movimentações 7 dias
        start_dt = datetime.utcnow() - timedelta(days=7)
        moves = s.execute(
            select(Auditoria.created_at)
            .where(Auditoria.acao == "UPDATE_STOCK", Auditoria.created_at >= start_dt)
        ).scalars().all()

        # agrupa por dia
        day_map = {d: 0 for d in days}
        for dt in moves:
            dd = dt.date()
            if dd in day_map:
                day_map[dd] += 1

        values = [day_map[d] for d in days]
        kpi_moves7 = sum(values)

    return render_page(
        DASH_BODY,
        title="Dashboard",
        labels=labels,
        values=values,
        kpi_total_produtos=kpi_total,
        kpi_low=kpi_low,
        kpi_moves7=kpi_moves7,
        using_postgres=USING_POSTGRES,
        is_admin=is_admin(),
        current_user=current_user,
    )


# ------------------------------------------------------------------------------
# Exports (Excel)
# ------------------------------------------------------------------------------

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


@app.get("/export/produtos")
@login_required
def export_produtos():
    with Session(ENGINE) as s:
        produtos = s.execute(select(Produto).order_by(Produto.id.asc())).scalars().all()
        for p in produtos:
            _ = p.categoria

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    ws.append(["ID", "Nome", "Categoria", "SKU", "Preço", "Quantidade", "Estoque mínimo"])

    for p in produtos:
        ws.append([
            p.id,
            p.nome,
            p.categoria.nome if p.categoria else "",
            p.sku or "",
            float(p.preco),
            p.quantidade,
            p.estoque_minimo,
        ])

    autosize(ws)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="produtos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/export/auditoria")
@login_required
def export_auditoria():
    with Session(ENGINE) as s:
        itens = s.execute(select(Auditoria).order_by(Auditoria.id.desc()).limit(200)).scalars().all()
        for a in itens:
            _ = a.usuario
            _ = a.produto

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    ws.append(["Data", "Ação", "Usuário", "Produto ID", "Detalhe", "IP"])

    for a in itens:
        ws.append([
            a.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            a.acao,
            a.usuario.email if a.usuario else "",
            a.produto_id or "",
            a.detalhe or "",
            a.ip or "",
        ])

    autosize(ws)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="auditoria.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ------------------------------------------------------------------------------
# Error pages
# ------------------------------------------------------------------------------

@app.errorhandler(403)
def forbidden(_):
    return "Acesso negado (403). Somente Admin.", 403


# ------------------------------------------------------------------------------
# Run
# ------------------------------------------------------------------------------

if __name__ == "__main__":
    # Para rede local: 0.0.0.0
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=True)
